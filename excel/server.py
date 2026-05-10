from flask import Flask, request, jsonify
import pandas as pd
from openpyxl import load_workbook
import requests

app = Flask(__name__)
def get_region_value(region: str) -> str:
    with open("region_values.txt", "r") as f:
        for line in f:
            parts = line.strip().split(",")

            print(parts[0])
            if len(parts) == 7 and parts[0] == region:
                return parts[1]+","+parts[2]+","+parts[3]+","+parts[4]+","+parts[5]+","+parts[6]
    return "UnknownRegion"
# === Excel: Create File ===
@app.route("/createExcel", methods=["POST"])
def create_excel():
    data = request.json
    filename = data.get("filename", "output.xlsx")
    columns = data.get("columns", [])
    rows = data.get("data", [])

    df = pd.DataFrame([rows], columns=columns) if rows else pd.DataFrame(columns=columns)
    df.to_excel(filename, index=False)

    return jsonify({"status": "success", "filename": filename})


# === Add Family of 4 ===
@app.route("/addFamilyOf4", methods=["POST"])
def add_family_of4():
    data = request.json
    filename = data.get("filename")
    row = data.get("row")
    region = data.get("region")
    print(region)
    print(filename)
    if not filename or not row or not region:
            return jsonify({"error": "Missing filename, row, or region"}), 400

        # Lookup region value from text file
    region_value = get_region_value(region)
    try:
        wb = load_workbook(filename)
        ws = wb.active
        # Determine next sequential number
        next_seq = ws.max_row   # 1-based row index

        # Prepend sequence number to row
        full_row = [next_seq] + row + [region_value]
        ws.append(full_row)
        wb.save(filename)
        return jsonify({"status": "success", "filename": filename})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# === Add Family of 2 ===
@app.route("/addFamilyOf2", methods=["POST"])
def add_family_of2():
    data = request.json
    filename = data.get("filename")
    row = data.get("row")

    try:
        wb = load_workbook(filename)
        ws = wb.active
       # Determine next sequential number
        next_seq = ws.max_row   # 1-based row index

        # Prepend sequence number to row
        full_row = [next_seq] + row
        ws.append(full_row)

        wb.save(filename)
        return jsonify({"status": "success", "filename": filename})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# === Add Subscriber ===
@app.route("/addSubscriber", methods=["POST"])
def add_subscriber():
    data = request.json
    filename = data.get("filename")
    row = data.get("row")

    try:
        wb = load_workbook(filename)
        ws = wb.active
        # Determine next sequential number
        next_seq = ws.max_row   # 1-based row index

        # Prepend sequence number to row
        full_row = [next_seq] + row
        ws.append(full_row)

        wb.save(filename)
        return jsonify({"status": "success", "filename": filename})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# === ServiceNow Ticket ===
@app.route("/ticket", methods=["POST"])
def create_ticket():
    data = request.json
    title = data.get("title")
    priority = data.get("priority")

    # Map priority to ServiceNow values
    priority_map = {"Low": "3", "Medium": "2", "High": "1"}
    sn_priority = priority_map.get(priority, "3")

    payload = {
        "short_description": title,
        "priority": sn_priority,
        "category": "software"
    }

    # Replace with your ServiceNow instance details
    SERVICENOW_INSTANCE = "https://your-instance.service-now.com"
    SERVICENOW_USER = "your-username"
    SERVICENOW_PASS = "your-password"

    try:
        response = requests.post(
            f"{SERVICENOW_INSTANCE}/api/now/table/incident",
            auth=(SERVICENOW_USER, SERVICENOW_PASS),
            headers={"Content-Type": "application/json"},
            json=payload
        )
        response.raise_for_status()
        result = response.json()
        return jsonify({"status": "success", "ticket": result.get("result", {})})
    except Exception as e:
        # If you don’t want real ServiceNow calls, you can mock here
        return jsonify({"status": "mock", "ticket": {"title": title, "priority": priority}})


if __name__ == "__main__":
    app.run(port=5000)
